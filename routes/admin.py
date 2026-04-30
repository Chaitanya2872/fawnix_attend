"""
Admin Routes
Administrative endpoints
"""

from flask import Blueprint, jsonify, Response, send_file
from middleware.auth_middleware import token_required
from middleware.admin_middleware import hr_or_devtester_required
from services import admin_service
from services import field_visit_service
import csv
from io import StringIO, BytesIO
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
import openpyxl
from openpyxl.styles import Font
from services.notification_service import (
    create_scheduled_notification,
    get_notification_candidates,
    get_scheduled_notifications,
    get_scheduled_notification_logs,
    send_push_notification_to_employee,
    trigger_scheduled_notification,
)
from services.attendance_exceptions_service import get_team_exceptions

from database.connection import get_db_connection, return_connection
from datetime import datetime, date, time
from flask import request

admin_bp = Blueprint('admin', __name__)


def serialize_row(row):
    result = {}
    for key, value in row.items():
        if isinstance(value, (datetime, date, time)):
            result[key] = value.isoformat()
        else:
            result[key] = value
    return result


def _devtester_super_admin_required(current_user):
    designation = (current_user.get("emp_designation") or "").strip().lower()
    return designation == "devtester"


def _resolve_emp_code_from_user_id(user_id):
    try:
        normalized_user_id = int(user_id)
    except (TypeError, ValueError):
        return None, "user_id must be a valid integer"

    if normalized_user_id <= 0:
        return None, "user_id must be greater than 0"

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute(
            """
            SELECT emp_code
            FROM users
            WHERE id = %s
            """,
            (normalized_user_id,),
        )
        row = cursor.fetchone()
        if not row or not row.get("emp_code"):
            return None, "No employee code found for the provided user_id"

        return row["emp_code"], None
    finally:
        cursor.close()
        return_connection(conn)


@admin_bp.route('/employees', methods=['GET'])
@token_required
@hr_or_devtester_required
def get_employees(current_user):
    """
    Get all employees
    Accessible only by HR and DevTester
    """
    employees = admin_service.get_all_employees()

    return jsonify({
        "success": True,
        "count": len(employees),
        "data": [serialize_row(emp) for emp in employees]
    }), 200


@admin_bp.route('/employees/report', methods=['GET'])
@token_required
@hr_or_devtester_required
def download_employees_report(current_user):
    """
    Download employee directory report (CSV, PDF, or XLSX).
    Query params: format=csv|pdf|xlsx
    """
    report_format = (request.args.get('format') or 'csv').lower()
    employees = admin_service.get_all_employees()

    headers = [
        "Employee Code",
        "Full Name",
        "Designation",
        "Grade",
        "Department",
        "Email",
        "Contact",
        "Manager Name",
        "Manager Email",
        "Manager Code",
        "Status",
    ]

    def _format_grade(value):
        raw = (value or "").strip()
        if not raw:
            return ""

        normalized = raw.upper()
        compact = normalized.replace(" ", "").replace("-", "").replace("_", "")

        if normalized == "NF" or compact == "NONFLEXIBLE":
            return "NF"
        if normalized == "F" or compact == "FLEXIBLE":
            return "F"
        if normalized == "M" or compact == "MODERATE":
            return "M"
        return raw

    rows = []
    for emp in employees:
        manager_name = emp.get("manager_name") or emp.get("emp_manager") or ""
        manager_email = emp.get("manager_email") or emp.get("manager_code") or ""
        status_value = emp.get("is_active")
        status_label = "Active" if status_value is True else "Inactive" if status_value is False else ""

        rows.append([
            emp.get("emp_code", ""),
            emp.get("emp_full_name", ""),
            emp.get("emp_designation", ""),
            _format_grade(emp.get("emp_grade")),
            emp.get("emp_department", ""),
            emp.get("emp_email", ""),
            emp.get("emp_contact", ""),
            manager_name,
            manager_email,
            emp.get("emp_manager", "") or "",
            status_label,
        ])

    today_str = datetime.utcnow().strftime("%Y-%m-%d")

    if report_format == 'pdf':
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=landscape(letter))
        width, height = landscape(letter)
        x_start = 0.4 * inch
        y = height - 0.6 * inch

        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(x_start, y, f"Employee Directory Report - {today_str}")
        y -= 0.35 * inch

        pdf.setFont("Helvetica-Bold", 8)
        col_widths = [0.9, 2.1, 1.4, 0.8, 1.5, 2.0, 1.2, 1.9, 2.0, 1.1, 0.9]
        col_widths = [w * inch for w in col_widths]

        def draw_row(values, y_pos, bold=False):
            pdf.setFont("Helvetica-Bold" if bold else "Helvetica", 7)
            x = x_start
            for value, width_col in zip(values, col_widths):
                text = str(value or "")
                if len(text) > 45:
                    text = text[:42] + "..."
                pdf.drawString(x, y_pos, text)
                x += width_col

        draw_row(headers, y, bold=True)
        y -= 0.22 * inch

        for row in rows:
            if y < 0.6 * inch:
                pdf.showPage()
                y = height - 0.6 * inch
                draw_row(headers, y, bold=True)
                y -= 0.22 * inch

            draw_row(row, y)
            y -= 0.18 * inch

        pdf.save()
        buffer.seek(0)
        filename = f"employees_{today_str}.pdf"
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )

    if report_format == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employees"
        ws.append(headers)

        for row in rows:
            ws.append(row)

        column_widths = [15, 26, 18, 10, 20, 28, 16, 24, 26, 16]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"employees_{today_str}.xlsx"
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)

    filename = f"employees_{today_str}.csv"
    response = Response(output.getvalue(), mimetype='text/csv')
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response


@admin_bp.route('/admins', methods=['POST'])
@token_required
@hr_or_devtester_required
def create_admin_user_route(current_user):
    """
    Add or promote an employee as admin.

    Super admin:
        - DevTester only
    """
    if not _devtester_super_admin_required(current_user):
        return jsonify({
            "success": False,
            "message": "Only DevTester can add admins"
        }), 403

    data = request.get_json() or {}
    emp_code = (data.get('emp_code') or '').strip()
    can_read = bool(data.get('can_read', True))
    can_write = bool(data.get('can_write', False))

    response, status_code = admin_service.create_admin_user(
        emp_code=emp_code,
        can_read=can_read,
        can_write=can_write
    )
    return jsonify(response), status_code


@admin_bp.route('/admins/<string:emp_code>/permissions', methods=['GET'])
@token_required
@hr_or_devtester_required
def get_admin_permissions_route(current_user, emp_code):
    """Get admin read/write permissions. DevTester only."""
    if not _devtester_super_admin_required(current_user):
        return jsonify({
            "success": False,
            "message": "Only DevTester can view admin permissions"
        }), 403

    response, status_code = admin_service.get_admin_permissions(emp_code)
    return jsonify(response), status_code


@admin_bp.route('/admins/<string:emp_code>/permissions', methods=['PUT'])
@token_required
@hr_or_devtester_required
def update_admin_permissions_route(current_user, emp_code):
    """Update admin read/write permissions. DevTester only."""
    if not _devtester_super_admin_required(current_user):
        return jsonify({
            "success": False,
            "message": "Only DevTester can update admin permissions"
        }), 403

    data = request.get_json() or {}
    can_read = data['can_read'] if 'can_read' in data else None
    can_write = data['can_write'] if 'can_write' in data else None

    if can_read is not None and not isinstance(can_read, bool):
        return jsonify({
            "success": False,
            "message": "can_read must be boolean"
        }), 400

    if can_write is not None and not isinstance(can_write, bool):
        return jsonify({
            "success": False,
            "message": "can_write must be boolean"
        }), 400

    response, status_code = admin_service.update_admin_permissions(
        emp_code=emp_code,
        can_read=can_read,
        can_write=can_write
    )
    return jsonify(response), status_code


@admin_bp.route('/attendance/status', methods=['GET'])
@token_required
@hr_or_devtester_required
def get_all_attendance_status(current_user):
    """
    Get current attendance status for all employees
    Admin only
    """
    response, status_code = admin_service.get_all_attendance_status()

    return jsonify(response), status_code

@admin_bp.route('/attendance/history', methods=['GET'])
@token_required
@hr_or_devtester_required
def get_all_attendance_history(current_user):
    """
    Get attendance history for all employees
    Optional query params:
    - limit: number of records (optional, legacy)
    - page: page number (optional)
    - page_size: page size (optional)
    - date: YYYY-MM-DD (optional)
    """
    limit = request.args.get('limit', type=int)
    page = request.args.get('page', type=int)
    page_size = request.args.get('page_size', type=int)
    date_str = request.args.get('date')
    target_date = None

    if date_str:
        try:
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({
                "success": False,
                "message": "Invalid date format. Use YYYY-MM-DD"
            }), 400

    response, status_code = admin_service.get_all_attendance_history(
        limit=limit,
        target_date=target_date,
        page=page,
        page_size=page_size
    )

    return jsonify(response), status_code

DAILY_ATTENDANCE_HEADERS = [
    "Employee ID",
    "Employee Name",
    "Clock In Time",
    "Clock Out Time",
    "Duration",
]

MONTHLY_ATTENDANCE_HEADERS = [
    "Employee ID",
    "Employee Name",
    "Working Days",
    "Total Hours",
    "Average Hours",
    "Overtime",
]


def _daily_attendance_report_payload(target_date: date):
    base_records = admin_service.get_attendance_report_base_data(target_date=target_date)
    return admin_service.build_daily_attendance_report_rows(base_records)


def _monthly_attendance_report_payload(month: int, year: int):
    base_records = admin_service.get_attendance_report_base_data(month=month, year=year)
    daily_rows = admin_service.build_daily_attendance_report_rows(base_records)
    return admin_service.build_monthly_attendance_report_rows(daily_rows)


def _styled_pdf_table(table_data, col_widths, header_color):
    table = Table(
        table_data,
        colWidths=col_widths,
        repeatRows=1,
        hAlign='LEFT',
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), header_color),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("ALIGN", (1, 1), (1, -1), "LEFT"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def _export_daily_report(report_format: str, target_date: date, daily_rows):
    date_str = target_date.strftime("%Y-%m-%d")
    base_filename = f"daily_attendance_report_{date_str}"

    if report_format == 'csv':
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(DAILY_ATTENDANCE_HEADERS)
        for row in daily_rows:
            writer.writerow([
                row.get('employee_id', ''),
                row.get('employee_name', ''),
                row.get('clock_in_display', ''),
                row.get('clock_out_display', ''),
                row.get('duration_display', ''),
            ])

        response = Response(output.getvalue(), mimetype='text/csv')
        response.headers['Content-Disposition'] = f'attachment; filename={base_filename}.csv'
        return response

    if report_format == 'xlsx':
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Daily Report"
        sheet.append(DAILY_ATTENDANCE_HEADERS)

        for column_index in range(1, len(DAILY_ATTENDANCE_HEADERS) + 1):
            sheet.cell(row=1, column=column_index).font = Font(bold=True)

        for row_index, row in enumerate(daily_rows, start=2):
            sheet.cell(row=row_index, column=1, value=row.get('employee_id', ''))
            sheet.cell(row=row_index, column=2, value=row.get('employee_name', ''))

            clock_in_cell = sheet.cell(row=row_index, column=3)
            if isinstance(row.get('clock_in'), datetime):
                clock_in_cell.value = row['clock_in']
                clock_in_cell.number_format = "hh:mm AM/PM"
            else:
                clock_in_cell.value = "Incomplete"

            clock_out_cell = sheet.cell(row=row_index, column=4)
            if isinstance(row.get('clock_out'), datetime):
                clock_out_cell.value = row['clock_out']
                clock_out_cell.number_format = "hh:mm AM/PM"
            else:
                clock_out_cell.value = "Incomplete"

            duration_cell = sheet.cell(
                row=row_index,
                column=5,
                value=f'=IF(OR(NOT(ISNUMBER(C{row_index})),NOT(ISNUMBER(D{row_index})),D{row_index}<C{row_index}),"Incomplete",D{row_index}-C{row_index})',
            )
            duration_cell.number_format = "[h]:mm"

        if daily_rows:
            last_data_row = len(daily_rows) + 1
            totals_row = last_data_row + 1
            sheet.cell(row=totals_row, column=1, value="Total Duration").font = Font(bold=True)
            sheet.cell(row=totals_row, column=5, value=f"=SUM(E2:E{last_data_row})").font = Font(bold=True)
            sheet.cell(row=totals_row, column=5).number_format = "[h]:mm"

        sheet.freeze_panes = "A2"
        column_widths = [16, 24, 16, 16, 14]
        for idx, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"{base_filename}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=0.45 * inch,
        rightMargin=0.45 * inch,
        topMargin=0.45 * inch,
        bottomMargin=0.45 * inch,
    )
    styles = getSampleStyleSheet()
    story = [Paragraph(f"Daily Attendance Report - {date_str}", styles["Title"]), Spacer(1, 0.2 * inch)]

    table_data = [DAILY_ATTENDANCE_HEADERS]
    if daily_rows:
        for row in daily_rows:
            table_data.append([
                row.get('employee_id', ''),
                row.get('employee_name', ''),
                row.get('clock_in_display', ''),
                row.get('clock_out_display', ''),
                row.get('duration_display', ''),
            ])
    else:
        table_data.append(["No records found", "", "", "", ""])

    story.append(
        _styled_pdf_table(
            table_data,
            col_widths=[1.3 * inch, 2.5 * inch, 1.7 * inch, 1.7 * inch, 1.4 * inch],
            header_color=colors.HexColor("#1F2937"),
        )
    )

    document.build(story)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"{base_filename}.pdf",
        mimetype='application/pdf',
    )


def _export_monthly_report(report_format: str, month: int, year: int, monthly_rows):
    base_filename = f"monthly_attendance_report_{year}_{month:02d}"

    if report_format == 'csv':
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(MONTHLY_ATTENDANCE_HEADERS)
        for row in monthly_rows:
            writer.writerow([
                row.get('employee_id', ''),
                row.get('employee_name', ''),
                row.get('total_working_days', 0),
                row.get('total_hours_display', ''),
                row.get('average_hours_display', ''),
                row.get('total_overtime_display', ''),
            ])

        response = Response(output.getvalue(), mimetype='text/csv')
        response.headers['Content-Disposition'] = f'attachment; filename={base_filename}.csv'
        return response

    if report_format == 'xlsx':
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Monthly Report"
        sheet.append(MONTHLY_ATTENDANCE_HEADERS)

        for column_index in range(1, len(MONTHLY_ATTENDANCE_HEADERS) + 1):
            sheet.cell(row=1, column=column_index).font = Font(bold=True)

        for row_index, row in enumerate(monthly_rows, start=2):
            sheet.cell(row=row_index, column=1, value=row.get('employee_id', ''))
            sheet.cell(row=row_index, column=2, value=row.get('employee_name', ''))
            sheet.cell(row=row_index, column=3, value=row.get('total_working_days', 0))
            sheet.cell(row=row_index, column=4, value=(row.get('total_hours_minutes', 0) or 0) / 1440)
            sheet.cell(row=row_index, column=5, value=f"=IF(C{row_index}=0,0,D{row_index}/C{row_index})")
            sheet.cell(row=row_index, column=6, value=(row.get('total_overtime_minutes', 0) or 0) / 1440)

            sheet.cell(row=row_index, column=4).number_format = "[h]:mm"
            sheet.cell(row=row_index, column=5).number_format = "[h]:mm"
            sheet.cell(row=row_index, column=6).number_format = "[h]:mm"

        if monthly_rows:
            last_data_row = len(monthly_rows) + 1
            totals_row = last_data_row + 1
            sheet.cell(row=totals_row, column=1, value="Totals").font = Font(bold=True)
            sheet.cell(row=totals_row, column=3, value=f"=SUM(C2:C{last_data_row})").font = Font(bold=True)
            sheet.cell(row=totals_row, column=4, value=f"=SUM(D2:D{last_data_row})").font = Font(bold=True)
            sheet.cell(row=totals_row, column=5, value=f"=IF(C{totals_row}=0,0,D{totals_row}/C{totals_row})").font = Font(bold=True)
            sheet.cell(row=totals_row, column=6, value=f"=SUM(F2:F{last_data_row})").font = Font(bold=True)
            sheet.cell(row=totals_row, column=4).number_format = "[h]:mm"
            sheet.cell(row=totals_row, column=5).number_format = "[h]:mm"
            sheet.cell(row=totals_row, column=6).number_format = "[h]:mm"

        sheet.freeze_panes = "A2"
        column_widths = [16, 24, 14, 18, 16, 14]
        for idx, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"{base_filename}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=0.45 * inch,
        rightMargin=0.45 * inch,
        topMargin=0.45 * inch,
        bottomMargin=0.45 * inch,
    )
    styles = getSampleStyleSheet()
    story = [Paragraph("Monthly Attendance Report", styles["Title"]), Spacer(1, 0.2 * inch)]

    table_data = [MONTHLY_ATTENDANCE_HEADERS]
    if monthly_rows:
        for row in monthly_rows:
            table_data.append([
                row.get('employee_id', ''),
                row.get('employee_name', ''),
                str(row.get('total_working_days', 0)),
                row.get('total_hours_display', ''),
                row.get('average_hours_display', ''),
                row.get('total_overtime_display', ''),
            ])
    else:
        table_data.append(["No records found", "", "", "", "", ""])

    story.append(
        _styled_pdf_table(
            table_data,
            col_widths=[1.3 * inch, 2.4 * inch, 1.3 * inch, 1.6 * inch, 1.6 * inch, 1.2 * inch],
            header_color=colors.HexColor("#0F766E"),
        )
    )

    document.build(story)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"{base_filename}.pdf",
        mimetype='application/pdf',
    )


def _download_attendance_report_by_type(report_type: str):
    report_format = (request.args.get('format') or 'csv').lower()

    if report_format not in {'csv', 'xlsx', 'pdf'}:
        return jsonify({
            "success": False,
            "message": "Invalid format. Use csv, xlsx, or pdf."
        }), 400

    if report_type == 'daily':
        date_str = (request.args.get('date') or '').strip()
        if not date_str:
            return jsonify({
                "success": False,
                "message": "date is required for daily report. Use YYYY-MM-DD."
            }), 400

        try:
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({
                "success": False,
                "message": "Invalid date format. Use YYYY-MM-DD."
            }), 400

        daily_rows = _daily_attendance_report_payload(target_date)
        return _export_daily_report(report_format, target_date, daily_rows)

    if report_type == 'monthly':
        month = request.args.get('month', type=int)
        year = request.args.get('year', type=int)

        if not month or month < 1 or month > 12:
            return jsonify({
                "success": False,
                "message": "Invalid month. Use 1-12."
            }), 400

        if not year or year < 2000:
            return jsonify({
                "success": False,
                "message": "Invalid year."
            }), 400

        monthly_rows = _monthly_attendance_report_payload(month, year)
        return _export_monthly_report(report_format, month, year, monthly_rows)

    return jsonify({
        "success": False,
        "message": "Invalid report_type. Use daily or monthly."
    }), 400


@admin_bp.route('/attendance/report', methods=['GET'])
@token_required
@hr_or_devtester_required
def download_attendance_report(current_user):
    """
    Download attendance report by report_type.

    Query params:
    - date (YYYY-MM-DD) for daily report
    - month (1-12) and year (YYYY) for monthly report
    - format=csv|xlsx|pdf
    - report_type=daily|monthly (optional)
    """
    report_type = (request.args.get('report_type') or '').strip().lower()
    if not report_type:
        report_type = 'daily' if request.args.get('date') else 'monthly'
    return _download_attendance_report_by_type(report_type)


@admin_bp.route('/attendance/report/daily', methods=['GET'])
@token_required
@hr_or_devtester_required
def download_daily_attendance_report(current_user):
    """Download only daily attendance report in CSV/XLSX/PDF (requires ?date=YYYY-MM-DD)."""
    return _download_attendance_report_by_type('daily')


@admin_bp.route('/attendance/report/monthly', methods=['GET'])
@token_required
@hr_or_devtester_required
def download_monthly_attendance_report(current_user):
    """Download only monthly attendance summary in CSV/XLSX/PDF (requires ?month & ?year)."""
    return _download_attendance_report_by_type('monthly')

@admin_bp.route('/field-visits/<int:field_visit_id>/tracking', methods=['GET'])
@token_required
@hr_or_devtester_required
def get_field_visit_tracking(current_user, field_visit_id):
    """
    Get tracking points for a specific field visit (admin only).
    """
    response, status_code = field_visit_service.get_tracking_history(field_visit_id)
    return jsonify(response), status_code

@admin_bp.route('/attendance/summary', methods=['GET'])
@token_required
@hr_or_devtester_required
def get_all_day_summary(current_user):
    """
    Get day summary for all employees
    Optional query param: ?date=YYYY-MM-DD
    """
    date_str = request.args.get('date')
    target_date = None

    if date_str:
        try:
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({
                "success": False,
                "message": "Invalid date format. Use YYYY-MM-DD"
            }), 400

    response, status_code = admin_service.get_all_day_summary(target_date)

    return jsonify(response), status_code


@admin_bp.route('/holidays', methods=['GET'])
@token_required
@hr_or_devtester_required
def get_admin_holidays(current_user):
    """
    Get admin holidays list with automatic Sunday/second-Saturday entries.

    Query params:
    - year: YYYY (optional, default current year)
    - month: 1-12 (optional)
    """
    year = request.args.get('year', type=int) or date.today().year
    month = request.args.get('month', type=int)

    response, status_code = admin_service.get_admin_holidays(year=year, month=month)
    return jsonify(response), status_code


@admin_bp.route('/holidays', methods=['POST'])
@token_required
@hr_or_devtester_required
def create_admin_holiday(current_user):
    """
    Create an organization holiday.

    Request body:
    {
      "holidayName": "Diwali",
      "date": "YYYY-MM-DD",
      "holidayType": "Public Holiday|Company Holiday|Optional Holiday|Weekend",
      "description": "Festival holiday",
      "status": "Active|Inactive"
    }
    """
    payload = request.get_json() or {}
    response, status_code = admin_service.create_admin_holiday(
        payload=payload,
        created_by_emp_code=current_user.get('emp_code')
    )
    return jsonify(response), status_code


@admin_bp.route('/calendar-summary', methods=['GET'])
@token_required
@hr_or_devtester_required
def get_calendar_summary(current_user):
    """
    Get month-wise calendar summary.

    Query params:
    - month: 1-12 (optional, default current month)
    - year: YYYY (optional, default current year)
    - department: optional
    - employee / emp_code: optional employee code
    """
    month = request.args.get('month', type=int) or date.today().month
    year = request.args.get('year', type=int) or date.today().year
    department = request.args.get('department')
    emp_code = request.args.get('employee') or request.args.get('emp_code')

    response, status_code = admin_service.get_calendar_summary(
        month=month,
        year=year,
        department=department,
        emp_code=emp_code
    )
    return jsonify(response), status_code


@admin_bp.route('/activities', methods=['GET'])
@token_required
@hr_or_devtester_required
def get_all_activities(current_user):
    """
    Get activities for all employees
    Optional query params:
    - limit: number of records (default: 100)
    - type: activity_type filter (optional)
    - include_tracking: true/false (default: true) for field visit tracking points
    - include_activity_tracking: true/false (default: true) for activity GPS points
    """
    limit = request.args.get('limit', default=100, type=int)
    activity_type = request.args.get('type')
    include_tracking = request.args.get('include_tracking', default='true')
    include_tracking = str(include_tracking).lower() in ['1', 'true', 'yes']
    include_activity_tracking = request.args.get('include_activity_tracking', default='true')
    include_activity_tracking = str(include_activity_tracking).lower() in ['1', 'true', 'yes']

    response, status_code = admin_service.get_all_activities(
        limit=limit,
        activity_type=activity_type,
        include_tracking=include_tracking,
        include_activity_tracking=include_activity_tracking
    )

    return jsonify(response), status_code


@admin_bp.route('/leaves', methods=['GET'])
@token_required
@hr_or_devtester_required
def get_all_leaves(current_user):
    """
    Get leave requests for all employees
    Optional query params:
    - limit: number of records (default: 100)
    - status: pending/approved/rejected/cancelled (optional)
    - emp_code: filter by employee code (optional)
    - from_date: YYYY-MM-DD (optional)
    - to_date: YYYY-MM-DD (optional)
    """
    limit = request.args.get('limit', default=100, type=int)
    status = request.args.get('status')
    emp_code = request.args.get('emp_code')
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')

    from_date = None
    to_date = None

    if from_date_str:
        try:
            from_date = datetime.strptime(from_date_str, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({
                "success": False,
                "message": "Invalid from_date format. Use YYYY-MM-DD"
            }), 400

    if to_date_str:
        try:
            to_date = datetime.strptime(to_date_str, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({
                "success": False,
                "message": "Invalid to_date format. Use YYYY-MM-DD"
            }), 400

    response, status_code = admin_service.get_all_leaves(
        limit=limit,
        status=status,
        emp_code=emp_code,
        from_date=from_date,
        to_date=to_date
    )

    return jsonify(response), status_code

@admin_bp.route('/overtime-records', methods=['GET'])
@token_required
@hr_or_devtester_required
def get_all_overtime_records(current_user):
    """
    Get overtime records for all employees

    Optional query params:
    - limit: number of records (default: 100)
    - status: eligible/requested/approved/rejected/expired/utilized (optional)
    - emp_code: filter by employee code (optional)
    - from_date: YYYY-MM-DD (optional)
    - to_date: YYYY-MM-DD (optional)
    """

    limit = request.args.get('limit', default=100, type=int)
    status = request.args.get('status')
    emp_code = request.args.get('emp_code')
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')

    from_date = None
    to_date = None

    if from_date_str:
        try:
            from_date = datetime.strptime(from_date_str, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({
                "success": False,
                "message": "Invalid from_date format. Use YYYY-MM-DD"
            }), 400

    if to_date_str:
        try:
            to_date = datetime.strptime(to_date_str, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({
                "success": False,
                "message": "Invalid to_date format. Use YYYY-MM-DD"
            }), 400

    response, status_code = admin_service.get_all_overtime_records(
        limit=limit,
        status=status,
        emp_code=emp_code,
        from_date=from_date,
        to_date=to_date
    )

    return jsonify(response), status_code


@admin_bp.route('/test-push', methods=['POST'])
@token_required
@hr_or_devtester_required
def send_test_push(current_user):
    """
    Send a manual test push notification to an employee device.

    Request body:
        {
            "emp_code": "2872"
        }
    or
        {
            "user_id": 4
        }
    Optional fields:
        - title
        - body
        - data
    """
    data = request.get_json() or {}

    emp_code = (data.get('emp_code') or '').strip()
    if not emp_code and data.get('user_id') is not None:
        emp_code, error_message = _resolve_emp_code_from_user_id(data.get('user_id'))
        if error_message:
            return jsonify({
                "success": False,
                "message": error_message
            }), 400

    if not emp_code:
        return jsonify({
            "success": False,
            "message": "emp_code or user_id is required"
        }), 400

    title = (data.get('title') or 'Test Notification').strip()
    body = (data.get('body') or 'This is a test push notification from the backend.').strip()
    payload = data.get('data') if isinstance(data.get('data'), dict) else {}
    payload.setdefault('type', 'test_notification')
    payload.setdefault('employee_id', emp_code)
    payload.setdefault('attendance_id', '0')
    payload.setdefault('status', 'test')
    payload.setdefault('timestamp', datetime.utcnow().isoformat())

    result = send_push_notification_to_employee(
        emp_code,
        title,
        body,
        payload
    )

    result.update({
        "target_emp_code": emp_code,
        "requested_by": current_user.get('emp_code')
    })
    status_code = 200 if result.get("success") else 400
    return jsonify(result), status_code


@admin_bp.route('/scheduled-notifications/trigger', methods=['POST'])
@token_required
@hr_or_devtester_required
def trigger_scheduled_notification_route(current_user):
    """
    Manually trigger a scheduled notification flow.

    Request body:
        {
            "notification_type": "attendance_reminder" | "lunch_reminder",
            "target_date": "YYYY-MM-DD"  // optional, defaults to today
        }
    """
    data = request.get_json() or {}
    notification_type = (data.get('notification_type') or '').strip().lower()
    target_date = None
    target_date_raw = (data.get('target_date') or '').strip()
    selected_emp_codes_raw = data.get('emp_codes')

    if not notification_type:
        return jsonify({
            "success": False,
            "message": "notification_type is required"
        }), 400

    if target_date_raw:
        try:
            target_date = datetime.strptime(target_date_raw, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({
                "success": False,
            "message": "Invalid target_date format. Use YYYY-MM-DD"
        }), 400

    if selected_emp_codes_raw is None:
        selected_emp_codes = []
    elif isinstance(selected_emp_codes_raw, list):
        selected_emp_codes = [
            str(emp_code).strip()
            for emp_code in selected_emp_codes_raw
            if str(emp_code).strip()
        ]
    else:
        return jsonify({
            "success": False,
            "message": "emp_codes must be an array of employee codes"
        }), 400

    result = trigger_scheduled_notification(
        notification_type,
        target_date=target_date,
        emp_codes=selected_emp_codes,
    )
    result["requested_by"] = current_user.get('emp_code')
    status_code = 200 if result.get("success") else 400
    return jsonify(result), status_code


@admin_bp.route('/scheduled-notifications/candidates', methods=['GET'])
@token_required
@hr_or_devtester_required
def get_scheduled_notification_candidates_route(current_user):
    """
    Get eligible employees for a scheduled notification flow.

    Query params:
        notification_type: attendance_reminder | lunch_reminder
        target_date: YYYY-MM-DD, optional
    """
    notification_type = (request.args.get('notification_type') or '').strip().lower()
    target_date_raw = (request.args.get('target_date') or '').strip()
    target_date = None

    if not notification_type:
        return jsonify({
            "success": False,
            "message": "notification_type is required"
        }), 400

    if target_date_raw:
        try:
            target_date = datetime.strptime(target_date_raw, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({
                "success": False,
                "message": "Invalid target_date format. Use YYYY-MM-DD"
            }), 400

    result = get_notification_candidates(notification_type, target_date=target_date)
    status_code = 200 if result.get("success") else 400
    return jsonify(result), status_code


@admin_bp.route('/scheduled-notifications', methods=['POST'])
@token_required
@hr_or_devtester_required
def create_scheduled_notification_route(current_user):
    """
    Create a custom scheduled push notification.

    Request body:
        {
            "title": "Scheduled Alert",     // optional
            "body": "Custom message",       // required
            "scheduled_date": "YYYY-MM-DD", // required
            "scheduled_time": "HH:MM"       // required
        }
    """
    data = request.get_json() or {}
    title = (data.get('title') or 'Scheduled Alert').strip()
    body = (data.get('body') or '').strip()
    scheduled_date_raw = (data.get('scheduled_date') or '').strip()
    scheduled_time_raw = (data.get('scheduled_time') or '').strip()

    if not body:
        return jsonify({
            "success": False,
            "message": "body is required"
        }), 400

    if not scheduled_date_raw or not scheduled_time_raw:
        return jsonify({
            "success": False,
            "message": "scheduled_date and scheduled_time are required"
        }), 400

    try:
        scheduled_date = datetime.strptime(scheduled_date_raw, "%Y-%m-%d").date()
        scheduled_time = datetime.strptime(scheduled_time_raw, "%H:%M").time()
    except ValueError:
        return jsonify({
            "success": False,
            "message": "Invalid scheduled_date or scheduled_time format"
        }), 400

    scheduled_for = datetime.combine(scheduled_date, scheduled_time)
    result = create_scheduled_notification(
        title=title,
        body=body,
        scheduled_for=scheduled_for,
        created_by_emp_code=current_user.get('emp_code'),
    )
    status_code = 201 if result.get("success") else 400
    return jsonify(result), status_code


@admin_bp.route('/scheduled-notifications', methods=['GET'])
@token_required
@hr_or_devtester_required
def get_scheduled_notifications_route(current_user):
    """
    Get recent scheduled notifications.

    Query params:
    - limit: 1-200, default 25
    - status: optional filter
    """
    limit = request.args.get('limit', default=25, type=int)
    status = (request.args.get('status') or '').strip().lower() or None

    rows = get_scheduled_notifications(limit=limit, status=status)
    return jsonify({
        "success": True,
        "count": len(rows),
        "data": rows
    }), 200


@admin_bp.route('/scheduled-notifications/logs', methods=['GET'])
@token_required
@hr_or_devtester_required
def get_scheduled_notification_logs_route(current_user):
    """
    Get recent scheduled notification log rows.

    Query params:
    - limit: 1-200, default 25
    - notification_type: optional filter
    """
    limit = request.args.get('limit', default=25, type=int)
    notification_type = (request.args.get('notification_type') or '').strip().lower() or None

    rows = get_scheduled_notification_logs(limit=limit, notification_type=notification_type)
    return jsonify({
        "success": True,
        "count": len(rows),
        "data": rows
    }), 200


@admin_bp.route('/team-exceptions', methods=['GET'])
@token_required
@hr_or_devtester_required
def admin_team_exceptions(current_user):
    """
    Get attendance exceptions for manager's team (Admin view)
    
    Authorization:
        - HR/DevTester only
        - Shows exceptions where current_user is the assigned manager
    
    Query Params:
        status: pending, approved, rejected (optional)
        type: late_arrival, early_leave (optional)
    
    Example:
        GET /api/admin/team-exceptions?status=pending
    
    Response:
        {
            "success": true,
            "data": {
                "exceptions": [...],
                "count": 2,
                "pending_count": 2
            }
        }
    """
    status = request.args.get('status')
    exception_type = request.args.get('type')
    
    result = get_team_exceptions(
        current_user['emp_code'],
        status,
        exception_type
    )
    
    return jsonify(result[0]), result[1]


@admin_bp.route('/late-arrivals', methods=['GET'])
@token_required
@hr_or_devtester_required
def admin_late_arrivals(current_user):
    """
    Get only late-arrival exceptions for admin/team views.

    Query Params:
        status: pending, approved, rejected, not_informed (optional)
    """
    status = request.args.get('status')
    result = get_team_exceptions(
        current_user['emp_code'],
        status,
        'late_arrival'
    )
    return jsonify(result[0]), result[1]


@admin_bp.route('/early-leaves', methods=['GET'])
@token_required
@hr_or_devtester_required
def admin_early_leaves(current_user):
    """
    Get only early-leave exceptions for admin/team views.

    Query Params:
        status: pending, approved, rejected, not_requested (optional)
    """
    status = request.args.get('status')
    result = get_team_exceptions(
        current_user['emp_code'],
        status,
        'early_leave'
    )
    return jsonify(result[0]), result[1]

